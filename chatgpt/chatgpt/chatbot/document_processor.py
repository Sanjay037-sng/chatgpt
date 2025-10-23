"""
Document processing utilities for extracting text from various file formats
"""
import os
import io
from typing import Optional, Tuple
from django.core.files.uploadedfile import UploadedFile

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None


class DocumentProcessor:
    """Utility class for processing different document types"""
    
    @staticmethod
    def extract_text_from_file(uploaded_file: UploadedFile) -> Tuple[str, str]:
        """
        Extract text content from uploaded file
        
        Args:
            uploaded_file: Django UploadedFile object
            
        Returns:
            Tuple of (extracted_text, file_type)
        """
        file_name = uploaded_file.name.lower()
        file_extension = os.path.splitext(file_name)[1].lower()
        
        try:
            if file_extension == '.pdf':
                return DocumentProcessor._extract_from_pdf(uploaded_file), 'pdf'
            elif file_extension in ['.docx', '.doc']:
                return DocumentProcessor._extract_from_docx(uploaded_file), 'docx'
            elif file_extension in ['.xlsx', '.xls']:
                return DocumentProcessor._extract_from_excel(uploaded_file), 'excel'
            elif file_extension in ['.txt', '.md']:
                return DocumentProcessor._extract_from_text(uploaded_file), 'text'
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                return DocumentProcessor._extract_from_image(uploaded_file), 'image'
            else:
                return f"Unsupported file type: {file_extension}", 'unsupported'
        except Exception as e:
            return f"Error processing file: {str(e)}", 'error'
    
    @staticmethod
    def _extract_from_pdf(uploaded_file: UploadedFile) -> str:
        """Extract text from PDF file"""
        if not PyPDF2:
            return "PyPDF2 library not installed. Cannot process PDF files."
        
        try:
            uploaded_file.seek(0)
            pdf_reader = PyPDF2.PdfReader(uploaded_file)
            text = ""
            
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text += page.extract_text() + "\n"
            
            return text.strip()
        except Exception as e:
            return f"Error reading PDF: {str(e)}"
    
    @staticmethod
    def _extract_from_docx(uploaded_file: UploadedFile) -> str:
        """Extract text from DOCX file"""
        if not DocxDocument:
            return "python-docx library not installed. Cannot process DOCX files."
        
        try:
            uploaded_file.seek(0)
            doc = DocxDocument(uploaded_file)
            text = ""
            
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
            return text.strip()
        except Exception as e:
            return f"Error reading DOCX: {str(e)}"
    
    @staticmethod
    def _extract_from_excel(uploaded_file: UploadedFile) -> str:
        """Extract text from Excel file"""
        if not openpyxl:
            return "openpyxl library not installed. Cannot process Excel files."
        
        try:
            uploaded_file.seek(0)
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            
            return text.strip()
        except Exception as e:
            return f"Error reading Excel: {str(e)}"
    
    @staticmethod
    def _extract_from_text(uploaded_file: UploadedFile) -> str:
        """Extract text from text file"""
        try:
            uploaded_file.seek(0)
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    uploaded_file.seek(0)
                    content = uploaded_file.read().decode(encoding)
                    return content.strip()
                except UnicodeDecodeError:
                    continue
            
            # If all encodings fail, use utf-8 with error handling
            uploaded_file.seek(0)
            content = uploaded_file.read().decode('utf-8', errors='replace')
            return content.strip()
        except Exception as e:
            return f"Error reading text file: {str(e)}"
    
    @staticmethod
    def _extract_from_image(uploaded_file: UploadedFile) -> str:
        """Extract text from image using OCR"""
        if not Image or not pytesseract:
            return "PIL and pytesseract libraries not installed. Cannot process images with OCR."
        
        try:
            uploaded_file.seek(0)
            image = Image.open(uploaded_file)
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Extract text using OCR
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            return f"Error processing image with OCR: {str(e)}"
    
    @staticmethod
    def get_supported_file_types() -> list:
        """Get list of supported file types"""
        supported_types = ['.txt', '.md']
        
        if PyPDF2:
            supported_types.append('.pdf')
        
        if DocxDocument:
            supported_types.extend(['.docx', '.doc'])
        
        if openpyxl:
            supported_types.extend(['.xlsx', '.xls'])
        
        if Image and pytesseract:
            supported_types.extend(['.jpg', '.jpeg', '.png', '.gif', '.bmp'])
        
        return supported_types
    
    @staticmethod
    def is_file_type_supported(file_name: str) -> bool:
        """Check if file type is supported"""
        file_extension = os.path.splitext(file_name.lower())[1]
        return file_extension in DocumentProcessor.get_supported_file_types()
